from openpyxl import Workbook,load_workbook 
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def read_multiple_sheets(file_path):
    excel_file=pd.ExcelFile(file_path)
    sheet_all_names=excel_file.sheet_names
    sheets={sheet:pd.read_excel(file_path,sheet_name=sheet) for sheet in sheet_all_names}
    return sheets

def write_multiple_sheets(file_name):
    wb=Workbook()
    sheet_name=['sheet1','sheet2','sheet3']
    data1={'Name':['aa','kk'],'age':[25,80]}
    data2={'city':['bb','yy'],'popo':[33,77]}
    data3={'pro':['oo','ll'],'price':[99,33]}
    dataframes=[pd.DataFrame(data1),pd.DataFrame(data2),pd.DataFrame(data3)]
    for idx,sheet_name in enumerate(sheet_name):
        ws=wb.create_sheet(sheet_name)
        for row in dataframe_to_rows(dataframes[idx],index=False,header=True):
            ws.append(row)
    del wb['Sheet']
    wb.save(file_name)

def create_and_save_excel(file_name):
    data={
        'name':['Alice','Bob','Charlie'],
        'age':[25,30,35],
        'city':['New York','Los Angeles','Chicago']
    }
    df=pd.DataFrame(data)
    df.to_excel(file_name,index=False)

#创建和保存excel文件
# create_and_save_excel('example.xlsx')
#读取多个sheet
# sheet_data=read_multiple_sheets('example.xlsx')
#写入多个sheet
write_multiple_sheets('newfile.xlsx')
